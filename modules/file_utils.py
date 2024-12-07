import os
import shutil
import logging
from typing import Dict, List, Tuple, Optional
from openpyxl.styles import PatternFill, Font, Border, Side, NamedStyle
import pandas as pd
from pandas import DataFrame

# Get a logger for this module
logger = logging.getLogger(__name__)

def copy_and_rename_files() -> None:
    """
    Copies specified files from their source locations, renames them, and places them in the current directory.
    Includes error handling for missing files and logs each step of the process.
    """
    if not os.path.exists('data'):
        os.mkdir('data')            # Create data directory if it does not already exist
    # Define the source and destination paths
    downloads_path = os.getenv("DOWNLOADS_PATH")
    sap_gui_path = os.getenv("SAP_GUI_PATH")
    base_directory = os.getenv("BASE_DIRECTORY")
    
    # Define the files to be copied and their new names
    paint_processed_file = 'Paint Processed.csv'
    export_file = 'paint_inventory.XLSX'
    
    try:
        # Copy and rename "Paint Processed.csv"
        src = os.path.join(downloads_path, paint_processed_file)
        dest = os.path.join(base_directory, r'data\paint_processed.csv')
        shutil.copy(src, dest)
        logger.info(f"Successfully copied '{paint_processed_file}' to '{dest}'.")

        # Remove the original file from Downloads
        os.remove(src)
        logger.info(f"Successfully removed '{paint_processed_file}' from {downloads_path}.")
        
    except FileNotFoundError as e:
        logger.error(f"File not found: {paint_processed_file} not found in {downloads_path}. Details: {e}")
    except Exception as e:
        logger.error(f"An error occurred while copying/renaming {paint_processed_file}: {e}", exc_info=True)

    try:
        # Copy and rename "paint_inventory.XLSX"
        src = os.path.join(sap_gui_path, export_file)
        dest = os.path.join(base_directory, 'data/paint_inventory.xlsx')
        shutil.copy(src, dest)
        logger.info(f"Successfully copied '{export_file}' to '{dest}'.")
        
    except FileNotFoundError as e:
        logger.error(f"File not found: {export_file} not found in {sap_gui_path}. Details: {e}")
    except Exception as e:
        logger.error(f"An error occurred while copying/renaming {export_file}: {e}", exc_info=True)


def read_shift_parameters(filename: str = 'shift_schedules.txt') -> Dict[str, object]:
    """
    Reads a text file to determine working hours, shifts active, and Saturday hours.
    Logs any issues with file access or invalid data formats.
    
    Parameters:
        filename (str): The path to the shift parameters file. Default is 'shift_schedules.txt'.
    
    Returns:
        Dict[str, object]: A dictionary containing shift parameters (e.g., working hours, active shifts).
    """
    parameters: Dict[str, object] = {}
    try:
        with open(filename, 'r') as file:
            lines = file.readlines()
        #logger.info(f"Successfully read shift parameters from '{filename}'.")

        for line in lines:
            if line.strip() and not line.strip().startswith('#'):
                try:
                    key, value = line.strip().split('=')
                    key = key.strip()
                    value = value.strip()
                    if key in ['first_shift', 'second_shift', 'include_saturday']:
                        parameters[key] = value.lower() in ('true', '1')
                    elif key in ['first_hours', 'second_hours']:
                        parameters[key] = int(value)
                except ValueError:
                    logger.warning(f"Skipping invalid line in '{filename}': {line.strip()}")
                
    except FileNotFoundError as e:
        logger.error(f"File not found: {filename}. Details: {e}")
    except Exception as e:
        logger.error(f"An error occurred while reading '{filename}': {e}", exc_info=True)

    return parameters


def get_work_shifts(parameters: Dict[str, object]) -> List[Tuple[int, int, int, int]]:
    """
    Uses parameters from 'read_shift_parameters' to record the working hours for each shift if applicable.
    Logs the shifts determined for first and second shifts.
    
    Parameters:
        parameters (Dict[str, object]): A dictionary containing shift configuration.
    
    Returns:
        List[Tuple[int, int, int, int]]: A list of tuples representing the working hours for each shift.
    """
    shifts: List[Tuple[int, int, int, int]] = []
    
    try:
        if parameters.get('first_shift', True):
            first_hours = int(parameters.get('first_hours', 8))
            if first_hours == 8:
                shifts.append((5, 0, 11, 0))  # First shift morning
                shifts.append((11, 30, 13, 30))  # First shift noon
                #logger.info("Configured first shift for 8 hours.")
            elif first_hours == 10:
                shifts.append((5, 0, 11, 0))  # First shift morning
                shifts.append((11, 30, 15, 30))  # First shift extended noon
                #logger.info("Configured first shift for 10 hours.")
        
        if parameters.get('second_shift', True):
            second_hours = int(parameters.get('second_hours', 8))
            if second_hours == 8:
                shifts.append((20, 30, 0, 30))  # Second shift evening
                shifts.append((1, 0, 5, 0))  # Second shift night
                #logger.info("Configured second shift for 8 hours.")
            elif second_hours == 10:
                shifts.append((18, 30, 0, 30))  # Second shift extended evening
                shifts.append((1, 0, 5, 0))  # Second shift night
                #logger.info("Configured second shift for 10 hours.")
    except Exception as e:
        logger.error(f"An error occurred while determining work shifts: {e}", exc_info=True)
    
    return shifts


def get_work_days(parameters: Dict[str, object]) -> List[int]:
    """
    Determines work days based on parameters returned from read_shift_parameters.
    Logs the work days configured.
    
    Parameters:
        parameters (Dict[str, object]): A dictionary containing shift configuration.
    
    Returns:
        List[int]: A list of integers representing workdays (0 = Monday, 6 = Sunday).
    """
    work_days: List[int] = [0, 1, 2, 3, 4]  # Default: Monday to Friday
    try:
        if parameters.get('include_saturday', False):
            work_days.append(5)  # Add Saturday to work days
            #logger.info("Saturday included as a work day.")
        else:
            #logger.info("Saturday not included as a work day.")
            pass
    except Exception as e:
        logger.error(f"An error occurred while determining work days: {e}", exc_info=True)

    return work_days

def save_to_excel(df: DataFrame, filename: str) -> None:
    """
    Applies specific formatting to Excel files for this project and saves them.
    
    Parameters:
        df (DataFrame): The DataFrame to save and format in Excel.
        filename (str): The path where the Excel file will be saved.
    """
    try:
        # Define styles
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        bold_font = Font(bold=True)
        border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')

        logger.info("Starting Excel write operation to '%s'.", filename)

        # Define writer
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        
        # Write DataFrame to Excel
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        logger.info("DataFrame written to Excel in sheet 'Sheet1'.")

        # Access the workbook and sheet
        workbook = writer.book
        sheet = workbook['Sheet1']

        # Format 'Last addtn to stock' column as date (assuming it's column H)
        for cell in sheet['H']:  # Adjust if 'Last addtn to stock' column changes
            if cell.row != 1:  # Skip the header row
                cell.style = date_style
        logger.info("Applied date formatting to 'Last addtn to stock' column.")

        # Apply borders to all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
        logger.info("Applied borders to all cells in the worksheet.")

        # Apply styles to header
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = bold_font
        logger.info("Formatted header row with bold font and gray fill.")

        # Adjust column widths based on cell content
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column letter
            for cell in column:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except Exception as e:
                    logger.warning("Error processing cell '%s' for column width adjustment: %s", cell.coordinate, e)
            adjusted_width = max_length + 2  # Add a little extra space
            sheet.column_dimensions[column_letter].width = adjusted_width
        logger.info("Adjusted column widths based on content.")

        # Save the formatted Excel file
        writer._save()
        logger.info("Successfully saved the Excel file to '%s'.", filename)

    except Exception as e:
        logger.error("An error occurred while saving the Excel file: %s", e, exc_info=True)

def move_to_sharepoint(
    source_dir: Optional[str] = None,
    dest_dir: Optional[str] = None,
    files_to_move: Optional[List[str]] = None
        ) -> None:
    """
    Moves specified files from the source directory to the destination SharePoint directory.

    Args:
        source_dir (Optional[str]): The source directory where the files are located.
                                    If None, defaults to '<current working directory>/data'.
        dest_dir (Optional[str]): The destination directory (SharePoint path) where the files should be moved.
                                  This parameter must be provided.
        files_to_move (Optional[List[str]]): List of file names to move. If None, defaults to 
                                             ['aged_load_inv.xlsx', 'aged_unload_inv.xlsx'].

    Raises:
        ValueError: If the destination directory is not provided.
        FileNotFoundError: If a specified file is not found in the source directory.
        Exception: For any other errors that occur during the file moving process.
    """
    try:
        # Set default values
        if source_dir is None:
            source_dir = os.path.join(os.getcwd(), 'data')
            logger.debug(f"Source directory not provided. Defaulting to {source_dir}.")

        if dest_dir is None:
            raise ValueError("Destination directory must be provided.")
        
        if files_to_move is None:
            files_to_move = ['aged_load_inv.xlsx', 'aged_unload_inv.xlsx', 'aged_8qi_inv.xlsx']
            logger.debug(f"Files to move not specified. Defaulting to {files_to_move}.")

        # Ensure source directory exists
        if not os.path.exists(source_dir):
            raise FileNotFoundError(f"Source directory '{source_dir}' does not exist.")

        # Ensure destination directory exists
        if not os.path.exists(dest_dir):
            raise FileNotFoundError(f"Destination directory '{dest_dir}' does not exist.")

        # Move each file
        for file_name in files_to_move:
            src = os.path.join(source_dir, file_name)
            dst = os.path.join(dest_dir, file_name)
            try:
                shutil.copy(src, dst)
                logger.info(f"Successfully moved '{file_name}' from '{source_dir}' to '{dest_dir}'.")
            except FileNotFoundError:
                logger.error(f"File '{file_name}' not found in source directory '{source_dir}'.", exc_info=True)
                raise
            except Exception as e:
                logger.error(f"Error moving file '{file_name}' from '{source_dir}' to '{dest_dir}': {e}", exc_info=True)
                raise

    except Exception as e:
        logger.critical(f"Critical error in move_to_sharepoint: {e}", exc_info=True)
        raise